from numpy import product, sort
import openpyxl


my_inventory_file = openpyxl.load_workbook("inventory.xlsx")
product_list = my_inventory_file["Sheet1"]

products_per_supplier  ={}
total_value_per_supplier ={}
products_with_inventory_less_than_10 = {}

for product_row in range(2, product_list.max_row + 1):
    supplier_name = product_list.cell(product_row, 4).value
    inventory = product_list.cell(product_row, 2).value
    price_of_product = product_list.cell(product_row, 3).value
    product_number = product_list.cell(product_row, 1).value
    inventory_price = product_list.cell(product_row, 5)

# EXERCISE NO.1
# calculation of number of products per supplier 
    if supplier_name in products_per_supplier:
        current_number_of_products = products_per_supplier.get(supplier_name)
        products_per_supplier[supplier_name] = current_number_of_products + 1
    else:
        products_per_supplier[supplier_name] = 1
#EXERCISE NO.2
# calculation of total value of inventory per supplier
    if supplier_name in total_value_per_supplier:
        current_total_value = total_value_per_supplier.get(supplier_name)
        total_value_per_supplier[supplier_name] = current_total_value + inventory * price_of_product
    else:
        total_value_per_supplier[supplier_name] = inventory * price_of_product

#EXERCISE NO.3
# logic with inventory less than 10
    if inventory < 10:
        products_with_inventory_less_than_10[product_number] = inventory


#EXERCISE NO.4
# add value for total inventory price
    inventory_price.value =  inventory * price_of_product
    my_inventory_file.save("updated_inventory_file.xlsx")


print(products_per_supplier)
print(total_value_per_supplier)
print(products_with_inventory_less_than_10)
print(inventory_price)


